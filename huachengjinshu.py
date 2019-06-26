#!/usr/bin/env python  
# _*_ coding:utf-8 _*_  
# @Author  : lusheng


from pylab import *
from xlrd import open_workbook
from xlrd import xldate_as_datetime
import datetime
from openpyxl import load_workbook
from pyecharts import Line,configure
import requests
from selenium import webdriver
import time
from wand.image import Image
from PIL import Image as mage
import pytesseract
import io
import re
import globalmap as gl

# gl._init()



def sign_in(username,password,sd,ed):
    url = 'http://www.hme01.com/information/'
    options = webdriver.ChromeOptions()
    prefs = {"profile.managed_default_content_settings.images": 2}
    options.add_experimental_option("prefs", prefs)
    options.add_argument('--headless')
    browser = webdriver.Chrome(chrome_options=options)
    browser.get(url)
    time.sleep(5)

    # browser.get(url)
    # js = "document.getElementById('txtUser_Pwd').style.display='block'"  # 编写JS语句
    # browser.execute_script(js)  # 执行JS
    browser.find_element_by_xpath('//*[@id="user_id"]').send_keys(username)
    # print(browser.find_element_by_xpath('//*[@id="txtUser_Pwd"]').is_displayed())
    browser.find_element_by_xpath('//*[@id="user_pass"]').send_keys(password)
    browser.find_element_by_xpath('//*[@id="login_form"]/p[3]/span[1]/input').click()
    time.sleep(2)
    # cookies = browser.get_cookies()
    # print(cookies)
    # cookiestext = ''
    # for cookie in cookies:
    #     cookiestext = cookiestext + cookie['name'] + '=' +cookie['value'] + ';'
    # print(cookiestext)
    # return cookiestext
    browser.get('http://www.hme01.com/information/data')
    browser.find_element_by_xpath('//*[@id="option1"]/option[2]').click() #锰产业
    time.sleep(1)
    browser.find_element_by_xpath('//*[@id="option2"]/option[2]').click() #金属锰
    time.sleep(1)
    browser.find_element_by_xpath('//*[@id="option3"]/option[3]').click() #电解锰
    time.sleep(1)
    browser.find_element_by_xpath('//*[@id="InformationType"]/option[3]').click() #出厂价
    time.sleep(1)
    browser.find_element_by_xpath('//*[@id="InformationArea"]').send_keys('吉首市')
    browser.find_element_by_xpath('//*[@id="InformationStandard"]').send_keys('99.7')
    browser.find_element_by_xpath('//*[@id="sd"]').send_keys(sd)
    browser.find_element_by_xpath('//*[@id="ed"]').send_keys(ed)
    browser.find_element_by_xpath('//*[@id="main"]/div[1]/table/tbody/tr[2]/td/input[3]').click()

    dibulan = browser.find_elements_by_xpath('//*[@id="main"]/table/tbody/tr[last()]/td/div//span')
    if len(dibulan) == 1:
        rows = browser.find_elements_by_xpath('//*[@id="main"]/table/tbody//tr[@onmouseover="this.className=\'th_on\'"]')
        for row in reversed(rows):
            price = row.find_element_by_xpath('./td[4]').text
            avgprice = row.find_element_by_xpath('./td[5]').text
            date = row.find_element_by_xpath('./td[7]').text
            print(price,avgprice,date)
            save_data(price, avgprice, date)
    else:
        dibulan2 = browser.find_elements_by_xpath('//*[@id="main"]/table/tbody/tr[22]/td/div//span')
        dibulan2[-1].find_element_by_xpath('./a').click()
        page = browser.find_element_by_xpath('//*[@id="main"]/table/tbody//tr/td/div//span[@class="current"]').text


        for i in reversed(range(int(page))):
            browser.get('http://www.hme01.com/information/data/page:%d/cate_f:41/cate_s:43/cate_t:87/type:147/s_date:%s/e_date:%s/area:吉首市/standard:48' % (i+1,sd,ed))
            rows = browser.find_elements_by_xpath('//*[@id="main"]/table/tbody//tr[@onmouseover="this.className=\'th_on\'"]')
            for row in reversed(rows):
                price = row.find_element_by_xpath('./td[4]').text
                avgprice = row.find_element_by_xpath('./td[5]').text
                date = row.find_element_by_xpath('./td[7]').text
                print(price,avgprice,date)
                save_data(price,avgprice,date)
    get_report(browser)
    browser.quit()
# def get_price_data(cookies):
    # data = {
    #     '_method': 'POST',
    #     'data[Information][cate_f]': '41',
    #     'data[Information][cate_s]': '43',
    #     'data[Information][cate_t]': '87',
    #     'data[Information][type]': '147',
    #     'data[Information][area]': '吉首市',
    #     'data[Information][standard]': '99.7',
    #     'data[Information][org_id]': '0',
    #     'data[Information][s_date]': '2019-02-01',
    #     'data[Information][e_date]': '2019-03-18',
    # }
    # headers = {
    #     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.75 Safari/537.36',
    #     'Cookie': cookies,
    #     'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    #     'Accept-Encoding': 'gzip, deflate',
    #     'Accept-Language': 'zh-CN,zh;q=0.9',
    #     'Cache-Control': 'max-age=0',
    #     'Connection': 'keep-alive',
    #     'Content-Type': 'application/x-www-form-urlencoded',
    #     'Host': 'www.hme01.com',
    #     'Upgrade-Insecure-Requests': '1',
    #     'Content-Length': '1000',
    # }
    # req = requests.post('http://www.hme01.com/information/data', data, headers=headers)
    # print(req.text)


def save_data(price, avgprice, date):
    excel_path = 'C:\\Users\\LUS\\Desktop\\周报材料\\周分析会议数据.xlsx'
    # 打开已经存在的表格并实例化，准备进行修改操作
    wb = load_workbook(excel_path)
    # print(wb.sheetnames)
    sheet = wb.get_sheet_by_name("电解锰片价格")
    # 查找需要修改的内容
    n_of_rows = sheet.max_row + 1
    n_of_cols = sheet.max_column
    haddate = []
    for i in range(1, n_of_rows):
        haddate.append(str(sheet.cell(row=i, column=1).value)[:10])
    # print(n_of_rows, n_of_cols, haddate[-5:])
    # 写入数据参数包括行号、列号、和值（其中参数不止这些）
    # sheet["A%d" % n_of_rows].value = indextitle[:-10]

    if date in haddate:
        pass
    else:
        sheet["A%d" % n_of_rows].value = date
        sheet["A%d" % n_of_rows].number_format = 'yyyy-mm-dd'
        sheet["B%d" % n_of_rows].value = price
        sheet["C%d" % n_of_rows].value = avgprice
    wb.save('C:\\Users\\LUS\\Desktop\\周报材料\\周分析会议数据.xlsx')



def make_picture():
    x_data = []
    y_data = []
    mpl.rcParams['font.sans-serif'] = ['SimHei']
    mpl.rcParams['axes.unicode_minus'] = False
    wb = open_workbook('C:\\Users\\LUS\\Desktop\\周报材料\\周分析会议数据.xlsx')
    s = wb.sheet_by_name(u'电解锰片价格')
    for row in range(s.nrows - 250, s.nrows):
        # print('the row is:',row )
        if type(s.cell_value(row, 0)) is str:
            date = s.cell_value(row, 0)
        else:
            date = xldate_as_datetime(s.cell_value(row, 0), 0)
            date = str(date.strftime('%Y-%m-%d'))
        # print(date)
        # values.append(date)
        x_data.append(date)
        y_data.append(s.cell_value(row, 2))
    x_data = [datetime.datetime.strptime(str(d), '%Y-%m-%d').date() for d in x_data]

    # 使用echarts绘图
    configure(output_image=True)
    line = Line('', background_color='white', title_text_size=25)
    attr = x_data
    v1 = y_data
    line.add('电解锰片价格', attr, v1, mark_line=['average'], is_label_show=False, is_smooth=True, line_width=3)
    line.render(path='C:\\Users\\LUS\Desktop\\周报材料\\锰片价格变化.jpeg')
    print('锰片价格作图完成')

def get_report(browser):
    browser.get('http://www.hme01.com/mn')
    title = browser.find_element_by_xpath('//*[@id="main"]/div[3]/div[6]/div/div[2]/ul/li[1]/a').get_attribute('title')
    link = browser.find_element_by_xpath('//*[@id="main"]/div[3]/div[6]/div/div[2]/ul/li[1]/a').get_attribute('href')
    print(title)
    browser.get(link)
    report_url = browser.find_element_by_xpath('//*[@id="main"]/div[2]/div[1]/div/div[3]/table/tbody/tr/td[1]/a').get_attribute('href')
    # report_url = 'dsds'
    response = requests.get(report_url)
    print(response)
    with open('C:\\Users\\LUS\Desktop\\周报材料\\' + '电解锰周评' + '.pdf','wb') as f:
        f.write(response.content)      # r.content -> requests中的二进制响应内容：以字节的方式访问请求响应体，对于非文本请求
        f.close()

    #pdf ocr方式提取文本
    # req_image = []
    # final_text = ''
    # with Image(filename=('C:\\Users\\LUS\Desktop\\周报材料\\' + title + '.pdf'), resolution=400) as img:
    #     with img.convert('jpeg') as converted:
    #         # converted.save(filename='image.jpeg')
    #         for img in converted.sequence:
    #             img_page = Image(image=img)
    #             req_image.append(img_page.make_blob('jpeg'))
    #
    # #为每个图像运行OCR，识别图像中的文本
    # for img in req_image:
    #     text = pytesseract.image_to_string(mage.open(io.BytesIO(img)), lang='chi_sim')
    #     final_text = final_text + text
    # # text = pytesseract.image_to_string(mage.open('image-0.jpeg'),lang='chi_sim')
    # final_text = final_text.replace(' ', '').replace('\n', '')
    # print(final_text)
    # mengpian_text = re.search(r'(?<=分析预测)(.+?)(?=\(原创)', final_text).group()
    # print(mengpian_text)
    #
    # f_mysteel = open('C:\\Users\\LUS\\Desktop\\周报材料\\华诚金属.txt', 'w', encoding='utf-8')
    # f_mysteel.write(mengpian_text)
    # f_mysteel.close()
    print('华诚金属网数据下载完成')

def huachengjinshu():
    today = datetime.datetime.now().strftime('%Y-%m-%d')
    strday = (datetime.datetime.now()-datetime.timedelta(days=15)).strftime("%Y-%m-%d")
    sign_in('sinometal', '20091228',strday,today)
    make_picture()
# get_price_data(cookies)


# huachengjinshu()


