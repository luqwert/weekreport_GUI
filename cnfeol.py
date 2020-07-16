#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# @Author  : Administrator.DESKTOP-4V3P1KOheng


from selenium import webdriver
import time
import re
from openpyxl import Workbook

def cnfeol(username,password,cookies):
    url = 'http://www.cnfeol.com/member/membersigninfrombox.aspx'
    # headers2 = {
    #     'Accept': 'text / html, application / xhtml + xml, application / xml;q = 0.9, image / webp, image / apng, * / *;q = 0.8',
    #     'Accept-Encoding':'gzip,deflate',
    #     'Accept-Language': 'zh-CN,zh;q = 0.9',
    #     'Connection': 'keep - alive',
    #     'Cookie': 'bdshare_firstime=1465884234301; ls_token=047ada6b-ce18-47ca-8ead-4ca01a9632f2; evercookie_png=047ada6b-ce18-47ca-8ead-4ca01a9632f2; evercookie_etag=047ada6b-ce18-47ca-8ead-4ca01a9632f2; evercookie_cache=047ada6b-ce18-47ca-8ead-4ca01a9632f2; cnfeol_mn=D4DF29036A2136C443B2919FB212EE01; cnfeol_smn=F6AD58D36D91CDBC5C420224F8B16EEF; ASP.NET_SessionId=eucn5p55ohnhyfm0jyka5eek; Hm_lvt_7d36fb642594f3133d486f18ce21e9fd=1546827489,1547088554,1547428481,1548036660; cnfeol_mn=; cnfeol_smn=; cnfeol_mkoc=false; cnfeol_mt_2015=40AE7CC5E559216E884991B0798EFF24EE6FE5E85F21D846639F5F1F77E0AA7D74BD1109697898835575B93FB7EEF1C91CE773EE0380698565D4F12AA5E0E9B788770A91AE3D74BB79EF160AFE9CA21A4FC0D9E00DF0C0911D887929BF071B8F3A6478176B8328D7D51BCAE2A8B3BF8AA0F117005CA01963BA7281BF3DB0BA288695F1407294325675C6DB2492099814B7EF416744DD620561F96DD6984C7575B5C34B7B30851329EE775ADCCD4D38180E8D3CE9AB4E912C4DC30C0366BA199AB8ADDBB91C86C3599B0E7D30CCACD51E772A6BAA1741B3DD23F20A02CE24572536BBDF79600CAB27B0A9C30B10B93D8D4DB511C9E09B5E32F532240BFEB12B3CD77A01876290D65144B7A7DED1E393905B44C67446311E60DA5734B10A9D53AA407817E2AB74055CD5182FC22153D4AE8DF893459A0AA46C692952B04E59C8C7B4BEAC352D98C24E3CB1EC92F0D2D2BB69FD9CF0B453F8CE76640F2B54A4858E6FEAB2D52BFB817C7F710C6C0A1CA3B9D9029A28AC959B4866B2722A635BFFDD; cnfeol_mt=bC2tT9Z9wi/dBNNEng0GvySuhBY74ZKx0x+ZcgnhqsYmfrB+GGMCeSV6zOMXz6OhJz2F4hbWFkoSNfOz/8GTcb8ut1Zou7B7h4H5xnDN2XrNauw1N5H9NRpGrgSAJBZbBVAEVtPh4/ADG7knE+psX/bEFkIG3s3pkviaJDicrwPXjLZluCuArNlnfhQlVjLE8oTbg35/HdpBtlzjrn4/9tcejK1do3MLMkjyr6kKPYiYLmX7UmqjflVl3yNY8XdG1pkmfN8SC/RVBXPDVDYISIfCt7Ykopcy3hQwHT53iIdnpXJg+zSwlBAdeW0EL22sNW8g5i+4deQ=; member=MTZjNzI1MzUzMTIzOTYxOTRjY2UwZjRlZjU0MWJhMzMyMDE5MDEyNjA5MzA0MERldmljZV9NZW1iZXJfMjk5NDA2; Hm_lpvt_7d36fb642594f3133d486f18ce21e9fd=1548379860',
    #     'Host':'www.cnfeol.com',
    #     'If-Modified-Since':'Fri, 25 Jan 2019 01:16:41 GMT',
    #     'If-None-Match': "86d7c7a04bb4d41:0",
    #     'Upgrade-Insecure-Requests': '1',
    #     'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.75 Safari/537.36'
    # }
    if cookies == '':
        headers = 'bdshare_firstime=1465884234301; ls_token=047ada6b-ce18-47ca-8ead-4ca01a9632f2; evercookie_png=047ada6b-ce18-47ca-8ead-4ca01a9632f2; evercookie_etag=047ada6b-ce18-47ca-8ead-4ca01a9632f2; evercookie_cache=047ada6b-ce18-47ca-8ead-4ca01a9632f2; cnfeol_mn=D4DF29036A2136C443B2919FB212EE01; cnfeol_smn=F6AD58D36D91CDBC5C420224F8B16EEF; ASP.NET_SessionId=r30jw255aehzn555olesvxno; Hm_lvt_7d36fb642594f3133d486f18ce21e9fd=1546481945,1546583487,1546827489,1547088554; cnfeol_mn=; cnfeol_smn=; cnfeol_mkoc=false; cnfeol_mt_2015=40AE7CC5E559216E884991B0798EFF24EE6FE5E85F21D846639F5F1F77E0AA7D74BD1109697898835575B93FB7EEF1C91CE773EE0380698565D4F12AA5E0E9B74B96D5EC936A445636B1BCCC64C728FA52CF2425A100D6CDEC9E2849F6897CE4A17815871D7091D63AFA69D535C2E8E46765689D1F7F8869F43E52CDE60AE7DE97D95D1A8C2B62FAE0FEBCD8A41F75C809F59DD2ABD875E3751874C75D58A125C73C6F00A61165FD5E2AABDA5B4BD9259EC8725034B2C51E8EF38BF358AA923D8EC679DEA66B1629B196665CE187B86D7B0B12C12D5A60FE0D0E74171C2850F2BF497FBD4225EB7D6AF015E9909FE4663208B907F7ABA6750CC8D80320DFE6511FC54523954FC6EDAF4935CA092056FB38A924601AE23C2633F60A41BB62F4D0C3F63A06187BDACE5AC84E4016806DC03C9D5FE8E75DE7E3D1BC8292F72854F033D4EB5564EC741F6A447BFA52F06B0EAE42D861FDE7886B2C9412137C70946E3442085503F078A8DDDD22CC80D6E67D662211D71BA24B0621FA16FAB01DB243; cnfeol_mt=bC2tT9Z9wi/dBNNEng0GvySuhBY74ZKx0x+ZcgnhqsYmfrB+GGMCeSV6zOMXz6OhJz2F4hbWFkoSNfOz/8GTcb8ut1Zou7B7h4H5xnDN2XrNauw1N5H9NRpGrgSAJBZbBVAEVtPh4/ADG7knE+psXxgSHXtARFSVUF1tClaR41VE2rt+mPYorwfbtUShJhBlhZJOQdejHT/KLNUd1IBmbcTOc2A60sLSb5lpt7iOn+SkxGsJg7NcTlVl3yNY8XdG7jxW1YDFipZv5lX6A9i07tRtIK2dDsEVo2A4INoWKGxnpXJg+zSwlBAdeW0EL22sNW8g5i+4deQ=; Hm_lpvt_7d36fb642594f3133d486f18ce21e9fd=1547100725'  # 没有备案过的浏览器，需要手机验证，无法手机验证，因此手动输入更新的cooki
    else:
        headers = cookies
    options = webdriver.ChromeOptions()
    options.add_argument(headers)
    prefs = {"profile.managed_default_content_settings.images": 2}
    options.add_experimental_option("prefs", prefs)
    # options.add_argument('--headless')
    browser = webdriver.Chrome(options=options)
    # time.sleep(3)
    browser.get('http://www.cnfeol.com/')
    input_username = browser.find_element_by_id('Signin_MemberName')
    input_username.send_keys(username)
    input_password = browser.find_element_by_id('Signin_MemberPassword')
    input_password.send_keys(password)
    signin = browser.find_element_by_id('Signin_Submit')
    signin.click()
    time.sleep(5)



    #进入锰矿/硅锰页面
    link = ''
    title = ''
    for i in range(1,5):
        browser.get('http://www.cnfeol.com/mengkuang/a-%d.aspx' % i)
        #获取分析报告列表
        lists = browser.find_elements_by_class_name('listbig')
        # print(lists)
        for list in lists:
            # print(list.get_attribute('title'))
            if '锰矿周评'in list.get_attribute('title'):
                link = list.get_attribute('href')
                title = list.get_attribute('title')
                print(link, title)
                break
        if link is not '':
            break

    #解析目标周报网页
    browser.get(link)
    time.sleep(3)
    #获取需要的部分内容
    page_text = browser.find_elements_by_xpath('//*[@id="contentdetail_info_detail"]//p')
    tables_mengkuang = browser.find_elements_by_xpath('//*[@id="contentdetail_info_detail"]//table')
    print(len(tables_mengkuang))
    if len(tables_mengkuang) >= 1:
        for table in tables_mengkuang:
            # print(table.find_element_by_xpath('./tbody/tr[1]/td[1]').text)
            if len(table.find_elements_by_xpath('./tbody//tr')) >= 1:
                if '港口' in table.find_element_by_xpath('./tbody/tr[1]/td[1]').text:
                    table1_tr = table.find_elements_by_xpath('./tbody//tr')
                    break
                else:
                    table1_tr = []
            else:
                table1_tr = []
        for table in tables_mengkuang:
            if len(table.find_elements_by_xpath('./tbody//tr')) >= 1:
                print(table.find_element_by_xpath('./tbody/tr[1]/td[1]').text)
                if '产品' in table.find_element_by_xpath('./tbody/tr[1]/td[1]').text:
                    table2_tr = table.find_elements_by_xpath('./tbody//tr')
                    print(len(table2_tr))
                    break
                else:
                    table2_tr = []
            else:
                table2_tr = []
    else:
        table1_tr = []
        table2_tr = []
        print('锰矿周分析页面没有表格')
    # table_content1 = browser.find_element_by_xpath('//*[@id="contentdetail_info_detail"]/table[1]').text
    # # print(table_content1)
    # # print(type(table_content1))
    # #构建表格自动填写的列表
    # table1_tr = browser.find_elements_by_xpath('//*[@id="contentdetail_info_detail"]/table[1]/tbody//tr')
    # # stock = []
    stock2 = []
    for tr in table1_tr:
        td = tr.find_elements_by_xpath('.//td')
        # print(td)
        # stock.append({'port': td[0].text, 'number1': td[1].text, 'number2': td[2].text, 'diff': td[3].text})
        tc = []
        for n in range(len(td)):
            tc.append(td[n].text)
        stock2.append(tc)
    print(stock2)
    wb = Workbook()
    ws = wb.active
    for li in stock2:
        ws.append(li)
    wb.save('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\cnfeol1.xlsx')

        # gl.set_value('stock', stock)

    # table_content2 = browser.find_element_by_xpath('//*[@id="contentdetail_info_detail"]/table[2]').text//*[@id="contentdetail_info_detail"]/table[3]/tbody
    # table2_tr = browser.find_elements_by_xpath('//*[@id="contentdetail_info_detail"]/table[2]/tbody//tr')
    mengkuang_price = []
    for tr in table2_tr:
        td = tr.find_elements_by_xpath('.//td')
        # print(td)
        # mengkuang_price.append({'type': td[0].text, 'spec': td[1].text, 'price1': td[2].text, 'price2': td[3].text, 'diff': td[4].text, 'port': td[5].text})
        tc = []
        for n in range(len(td)):
            tc.append(td[n].text)
        mengkuang_price.append(tc)

    print(mengkuang_price)
    wb = Workbook()
    ws = wb.active
    for li in mengkuang_price:
        ws.append(li)
    wb.save('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\cnfeol2.xlsx')

    # print(page_text)
    # text = title + '\n'
    text = ''
    for paragraph in page_text:
        text = text + paragraph.text + '\n'
    print(text)
    text = text.replace('\n', '')
    text1 = re.search(r'(?<=：)(.+?)(?=下游市场|硅锰方面|硅锰：|硅锰)', text).group() + '\r\n' + re.search(r'(?<=锰矿市场：)(.+?)(?=\s免责声明|免责声明|2020年铁合金行业)', text).group()
    print(text1)
    # print(table_content1)
    # print(table_content2)
    #保存获得的内容
    f_mengkuang = open('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\锰矿.txt', 'w', encoding='utf-8')
    f_mengkuang.write(text1)
    # f_mengkuang.write(table_content1 + '\n\n')
    # f_mengkuang.write(table_content2)
    f_mengkuang.close()

    #获取硅锰分析周报
    #获取分析报告列表
    link = ''
    title = ''
    for i in range(1,5):
        browser.get('http://www.cnfeol.com/guimeng/a-%d.aspx' % i)
        #获取分析报告列表
        lists = browser.find_elements_by_class_name('listbig')
        for list in lists:
            # print(list.get_attribute('title'))
            if '硅锰周评'in list.get_attribute('title'):
                link = list.get_attribute('href')
                title = list.get_attribute('title')
                print(link, title)
                break
        if link is not '':
            break


    #解析目标周报网页
    browser.get(link)
    time.sleep(5)

    #获取需要的部分内容
    page_text = browser.find_elements_by_xpath('//*[@id="contentdetail_info_detail"]//*')
    # print(page_text)
    # text = title + '\n'
    text = ''
    for paragraph in page_text:
        text = text + paragraph.text + '\n'

    # table_content3 = browser.find_element_by_xpath('//*[@id="contentdetail_info_detail"]/table').text
    # print(table_content3)
    text = text.replace('\n', '')
    # print(text)
    text2 = re.search(r'(?<=：|:)(.+?)(?=一、)',text).group() +'\r\n'+ re.search(r'(?<=本周总结：)(.+?)(?=四、趋势|四、预测)',text).group() +'\r\n'+ re.search(r'(?<=趋势：|预测：)(.+?)(?=免责声明|2020年铁合金行业)',text).group()

    print(text2)
    table3 = browser.find_elements_by_xpath('//*[@id="contentdetail_info_detail"]//table')
    print('页面有%d个表格' % len(table3))
    if len(table3) >= 2:
        for table in table3:
            # print(table.find_element_by_xpath('./tbody/tr[2]/td[1]').text)
            if '硅锰' in table.find_element_by_xpath('./tbody/tr[2]/td[1]').text:
                table3_tr = table.find_elements_by_xpath('./tbody//tr')
                break
            elif '钢厂' in table.find_element_by_xpath('./tbody/tr[1]/td[1]').text:
                table3_tr = table.find_elements_by_xpath('./tbody//tr')
                break
            elif '产品' in table.find_element_by_xpath('./tbody/tr[1]/td[1]').text:
                table3_tr = table.find_elements_by_xpath('./tbody//tr')
                break
            else:
                table3_tr = []
    elif len(table3) == 1:
        table3_tr = table3[0].find_elements_by_xpath('./tbody//tr')
    else:
        print('硅锰周分析页面没有表格')
        table3_tr = []
    # table3_tr = browser.find_elements_by_xpath('//*[@id="contentdetail_info_detail"]/table/tbody//tr')
    # print(len(table3_tr))
    guimeng_price = []
    for tr in table3_tr:
        td = tr.find_elements_by_xpath('.//td')
        tc = []
        for n in range(len(td)):
            tc.append(td[n].text)
        guimeng_price.append(tc)
    print(guimeng_price)
    wb = Workbook()
    ws = wb.active
    for li in guimeng_price:
        ws.append(li)
    wb.save('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\cnfeol3.xlsx')
    #关闭浏览器
    browser.quit()

    #保存获得的内容
    f_guimeng = open('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\硅锰.txt', 'w', encoding='utf-8')
    f_guimeng.write(text2)
    #关闭文件
    f_guimeng.close()

#
# cookies = 'bdshare_firstime=1465884234301; ls_token=047ada6b-ce18-47ca-8ead-4ca01a9632f2; evercookie_png=047ada6b-ce18-47ca-8ead-4ca01a9632f2; evercookie_etag=047ada6b-ce18-47ca-8ead-4ca01a9632f2; evercookie_cache=047ada6b-ce18-47ca-8ead-4ca01a9632f2; ASP.NET_SessionId=l33i4ganx2lmrdrzu5uadumw; cnfeol_mn=D4DF29036A2136C443B2919FB212EE01; cnfeol_smn=F6AD58D36D91CDBC5C420224F8B16EEF; cnfeol_mn=; cnfeol_smn=; Hm_lvt_7d36fb642594f3133d486f18ce21e9fd=1566786622; cnfeol_mkoc=false; cnfeol_mt_2015=40AE7CC5E559216E884991B0798EFF24EE6FE5E85F21D8467196F4694031AC4BB323ECF025369B9252EE8B0EA1E9C40BACD8453339B8171B8CB6181A1540938F1FBB112DE13A4EAB9DF426C00C70CA15F0C30E77F51B479F5A17D4528B4FD9474D21080CCD99BEDABA7568B988AA03AA0C64F8C07F253225636AC18553EDC76EC83F2674710523F3FEBF4ECDF4298ADBEE071E88790A7659B8D25E66E79EC86AA3C09ED6A484165E4119B3FEFFAB46FC534F74027FE4E8BEFAA247915CC394CE1E97054DFF04940F78B930CD66BABE6090A07A403B67D9093A6A4E95D536257CD82B6FC757ABC8C54E832373742EF722EF04547967140F4F19BF28502064BEA1A41638EB8E406825DB6526E087E85B98EB59D462C37C5C569BA049F6FABEE2336E6666A2725E34E7B7DBA887EA12408044570B08DE577FDF84585CFDA5315735AE6FBC08D9106C2D508D7CEE75998AF04EB8D6AB36DFA9FBD6B552182A65297F3B117DCCA8AC74B1296BCCCC6FBFB93EE6EAB2CC4AC3E7705F672F5C3EF3AF04; cnfeol_mt=bC2tT9Z9wi/dBNNEng0GvySuhBY74ZKx0x+ZcgnhqsYmfrB+GGMCeSV6zOMXz6OhJz2F4hbWFkoSNfOz/8GTcb8ut1Zou7B7h4H5xnDN2XrNauw1N5H9NRpGrgSAJBZbBVAEVtPh4/ADG7knE+psX86mvKXVq78dDIRr8gf0L0rSwjrA51RMC7wneXWhJws0r2bSIdDLThkUoGphrGIF2fBny5pUzzDu0j/w1x1YY46lnFRKS3PWWFVl3yNY8XdGMu5d5lRFZvrVg+5L5Lpfm7zioEUYAHNro2A4INoWKGxnpXJg+zSwlBAdeW0EL22sNW8g5i+4deQ=; Hm_lpvt_7d36fb642594f3133d486f18ce21e9fd=1567995523'  # 没有备案过的浏览器，需要手机验证，无法手机验证，因此手动输入更新的cooki
# cnfeol('sinometal', '20091228', cookies)

