#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# @Author  : lusheng

from openpyxl import load_workbook


excel_path = 'C:\\Users\\LUS\\Desktop\\分析.xlsx'

# 打开已经存在的表格并实例化，准备进行修改操作
wb = load_workbook(excel_path)
# print(wb.sheetnames)
sheet = wb.get_sheet_by_name("资产负债表")
# 查找需要修改的内容
n_of_rows = sheet.max_row +1
n_of_cols = sheet.max_column +1
for i in range(2,n_of_rows):
    for j in range(1,n_of_cols):
        if (type(sheet.cell(row=i, column=j).value) is str) and ('亿' in sheet.cell(row=i, column=j).value):
            print(i, j)
            print(sheet.cell(row=i, column=j).value)
            # print(str(int(float(sheet.cell(row=i, column=j).value[:-1]) * 10000)) + '万')
            c = str(int(float(sheet.cell(row=i, column=j).value[:-1]) * 10000)) + '万'
            sheet.cell(row=i, column=j).value = c
            print(sheet.cell(row=i, column=j).value)
wb.save('C:\\Users\\LUS\\Desktop\\分析2.xlsx')

sheet2 = wb.get_sheet_by_name("Sheet2")
# 查找需要修改的内容
n_of_rows = sheet2.max_row +1
n_of_cols = sheet2.max_column +1
for i in range(2,n_of_rows):
    for j in range(1,n_of_cols):
        if (type(sheet2.cell(row=i, column=j).value) is str) and ('亿' in sheet2.cell(row=i, column=j).value):
            print(i, j)
            print(sheet2.cell(row=i, column=j).value)
            # print(str(int(float(sheet.cell(row=i, column=j).value[:-1]) * 10000)) + '万')
            c = str(int(float(sheet2.cell(row=i, column=j).value[:-1]) * 10000)) + '万'
            sheet2.cell(row=i, column=j).value = c
            print(sheet2.cell(row=i, column=j).value)
wb.save('C:\\Users\\LUS\\Desktop\\分析2.xlsx')
sheet3 = wb.get_sheet_by_name("Sheet3")
# 查找需要修改的内容
n_of_rows = sheet3.max_row +1
n_of_cols = sheet3.max_column +1
for i in range(2,n_of_rows):
    for j in range(1,n_of_cols):
        if (type(sheet3.cell(row=i, column=j).value) is str) and ('亿' in sheet3.cell(row=i, column=j).value):
            print(i, j)
            print(sheet3.cell(row=i, column=j).value)
            # print(str(int(float(sheet.cell(row=i, column=j).value[:-1]) * 10000)) + '万')
            c = str(int(float(sheet3.cell(row=i, column=j).value[:-1]) * 10000)) + '万'
            sheet3.cell(row=i, column=j).value = c
            print(sheet3.cell(row=i, column=j).value)
wb.save('C:\\Users\\LUS\\Desktop\\分析2.xlsx')
