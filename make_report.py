#!/usr/bin/env python
# _*_ coding:utf-8 _*_
# @Author  : Administrator.DESKTOP-4V3P1KOheng

from openpyxl import load_workbook
from docx import shared
from docxtpl import DocxTemplate, RichText, InlineImage
import os


def make_report():
    tpl = DocxTemplate('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\周报模板.docx')

    if os.path.exists('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\各产品周市场分析.docx'):
        os.remove('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\各产品周市场分析.docx')

    # rt = RichText('an exemple of ')
    # rt.add('a rich text', style='')
    # rt.add('some violet', color='#ff00ff')
    image1 = InlineImage(tpl,'C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\普氏指数.jpeg',width=shared.Cm(16))
    image2 = InlineImage(tpl,'C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\废钢指数近一年变化.png',width=shared.Cm(16))
    image3 = InlineImage(tpl,'C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\各地废钢市场价格.png',width=shared.Cm(16))
    image4 = InlineImage(tpl,'C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\锰片价格变化.jpeg',width=shared.Cm(15))

    excel_path = 'C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\周分析会议数据.xlsx'
    wb = load_workbook(excel_path)
    ws = wb.get_sheet_by_name("普氏、MYSTEEL指数")
    date1 = str(ws['A%d' % ws.max_row].value)[:10]
    date1_62 = str(ws['B%d' % ws.max_row].value)
    date1_58 = str(ws['C%d' % ws.max_row].value)
    date2 = str(ws['A%d' % (ws.max_row - 5)].value)[:10]
    date2_62 = str(ws['B%d' % (ws.max_row - 5)].value)
    date2_58 = str(ws['C%d' % (ws.max_row - 5)].value)
    wb.close()
    if float(date1_62) - float(date2_62) >= 0:
        updown_62 = '上涨'
        diff_62 = round(float(date1_62) - float(date2_62),2)
    else:
        updown_62 = '下跌'
        diff_62 = round(-(float(date1_62) - float(date2_62)),2)

    if float(date1_58) - float(date2_58) >= 0:
        updown_58 = '上涨'
        diff_58 = round(float(date1_58) - float(date2_58),2)
    else:
        updown_58 = '下跌'
        diff_58 = round(-(float(date1_58) - float(date2_58)),2)

    mysteeltext = []
    with open('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\mysteel.txt', 'r', encoding='utf-8') as f_mysteel:
        for line in f_mysteel:
            mysteeltext.append(line.strip('\n').split(','))
        # print(mysteeltext)
        kucun_text = mysteeltext[0][0]
        print(kucun_text)
        kaigong_text = mysteeltext[2][0]
        print(kaigong_text)
        fenxi_text = mysteeltext[4][0] + mysteeltext[6][0]
        print(fenxi_text)
        haiyunfei_text = mysteeltext[10][0]
        print(haiyunfei_text)
        feigang_text = mysteeltext[8][0]
        print(feigang_text)


    with open('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\锰矿.txt', 'r', encoding='utf-8') as f_mengkuang:
        mengkuang_text = f_mengkuang.read()
    with open('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\硅锰.txt', 'r', encoding='utf-8') as f_guimeng:
        guimeng_text = f_guimeng.read()

    stock2 = []
    line = []
    wb = load_workbook('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\cnfeol1.xlsx')
    ws = wb.active
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            line.append(ws.cell(row=(row + 1), column=(col + 1)).value)
        stock2.append(line)
        line = []
    wb.close()
    print(stock2)
    menggkuang_kucun = stock2[-1][-2]
    print(menggkuang_kucun)
    if float(stock2[-1][-2]) - float(stock2[-1][-3]) > 0:
        updown_mengkuang = '增加'
        # diff_mengkuang = str(round(-(float(stock2[-1][-2]) - float(stock2[-1][-3])))) + '吨'
        diff_mengkuang = str(stock2[-1][-1]) + '吨'
    elif float(stock2[-1][-2]) - float(stock2[-1][-3]) < 0:
        updown_mengkuang = '减少'
        # print(stock2[-1][-1][1:])
        diff_mengkuang = str(stock2[-1][-1])[1:] + '吨'
    elif float(stock2[-1][-2]) - float(stock2[-1][-3]) == 0:
        updown_mengkuang = '不变'
        diff_mengkuang = ''

    mengkuang_price = []
    line3 = []
    wb = load_workbook('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\cnfeol2.xlsx')
    ws = wb.active
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            line3.append(ws.cell(row=(row + 1), column=(col + 1)).value)
        mengkuang_price.append(line3)
        line3 = []
    wb.close()
    print(mengkuang_price)

    guimeng_price = []
    line4 = []
    wb = load_workbook('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\cnfeol3.xlsx')
    ws = wb.active
    for row in range(ws.max_row):
        for col in range(ws.max_column):
            line4.append(ws.cell(row=(row + 1), column=(col + 1)).value)
        guimeng_price.append(line4)
        line4 = []
    wb.close()
    print(guimeng_price)

    with open('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\周报材料\\华诚金属.txt', 'r', encoding='utf-8') as f_mengpian:
        mengpian_text = f_mengpian.read()

    #需要传入的数据
    context = {
        # 文本和数字
        'date1': date1,
        'date2': date2,
        'date1_62': date1_62,
        'updown_62': updown_62,
        'diff_62': diff_62,
        'date1_58': date1_58,
        'updown_58': updown_58,
        'diff_58': diff_58,
        'kucun_text': kucun_text,
        'kaigong_text': kaigong_text,
        'fenxi_text': fenxi_text,
        'haiyunfei_text': haiyunfei_text,
        'feigang_text': feigang_text,
        'menggkuang_kucun': menggkuang_kucun,
        'updown_mengkuang': updown_mengkuang,
        'diff_mengkuang': diff_mengkuang,
        'mengkuang_text': mengkuang_text,
        'guimeng_text': guimeng_text,
        'mengpian_text': mengpian_text,
        # 'mengpian_text2': mengpian_text2,
        #表格
        # 'stock': stock, #表格，传入参数为list嵌套字典
        'stock2': stock2,
        'mengkuang_price': mengkuang_price,
        'guimeng_price': guimeng_price,

        #图片
        'image1': image1,
        'image2': image2,
        'image3': image3,
        'image4': image4,
    }
    print(context)

    tpl.render(context)
    tpl.save('C:\\Users\\Administrator.DESKTOP-4V3P1KO\\Desktop\\各产品周市场分析.docx')
    print('周报已生成完毕')
    # weekreportsend.send_report()

# make_report()
